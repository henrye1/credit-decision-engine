import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


def normalize_rule(raw_rule: Dict[str, Any]) -> Dict[str, Any]:
    """
    Normalize rule fields to the standard structure.
    Handles various column name variations (e.g. 'Phase' vs 'phase', 'RuleID' vs 'rule_id').
    """
    def get_field(key_variants: List[str], default: Any = None) -> Any:
        for key in key_variants:
            if key in raw_rule and raw_rule[key] is not None and str(raw_rule[key]).strip():
                return raw_rule[key]
        return default

    return {
        "phase": get_field(["Phase", "phase", "PHASE"], ""),
        "rule_id": get_field(["RuleID", "rule_id", "RuleId", "ID"], ""),
        "description": get_field(["Description", "description", "DESCRIPTION"], ""),
        "action": get_field(["Action", "action", "ACTION", "Decision"], "Decline"),
        "disputable": get_field(["Disputable", "disputable", "DISPUTABLE"], "No").lower() == "yes",
        "rule_applied": get_field(["Rule Applied", "rule_applied", "RuleApplied"], "In_decision_engine"),
    }


def parse_workbook(workbook_path: str | Path) -> Dict[str, Any]:
    path = Path(workbook_path)
    workbook = load_workbook(filename=path, data_only=True)

    sheets: List[Dict[str, Any]] = []
    rules: List[Dict[str, Any]] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        for row in rows[1:]:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            record = {header: (row[index] if index < len(row) else None) for index, header in enumerate(headers)}
            normalized = normalize_rule(record)
            normalized["sheet_name"] = sheet.title
            rules.append(normalized)

        sheets.append({"name": sheet.title, "row_count": len(rows) - 1})

    return {"workbook": path.name, "sheet_count": len(sheets), "sheets": sheets, "rules": rules}


def save_rules_config(workbook_path: str | Path, output_path: str | Path) -> Dict[str, Any]:
    parsed = parse_workbook(workbook_path)
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(parsed, indent=2), encoding="utf-8")
    return parsed
